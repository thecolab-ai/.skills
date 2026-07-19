"""Parse Treasury publication resources and fiscal forecast workbooks."""

from __future__ import annotations

import io
import re
from html import unescape
from urllib.parse import urljoin, urlparse


class ComparisonError(ValueError):
    """The requested publications cannot be compared without guessing."""


def clean(value: str) -> str:
    return " ".join(unescape(re.sub(r"<[^>]+>", " ", value)).split())


def _normalise(value: object) -> str:
    return " ".join(re.findall(r"[a-z0-9]+", str(value or "").casefold()))


def parse_resources(html: str, source_url: str, retrieved_at: str) -> list[dict[str, str]]:
    host = urlparse(source_url).hostname
    rows = []
    for href, body in re.findall(r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', html, re.I | re.S):
        url = urljoin(source_url, unescape(href))
        title = clean(body)
        parsed = urlparse(url)
        path = parsed.path.lower()
        if parsed.hostname != host or not title:
            continue
        if parsed.fragment or url.rstrip("/") == source_url.rstrip("/"):
            continue
        if _normalise(title) in {"skip to main content", "skip to navigation"}:
            continue
        filename = path.rsplit("/", 1)[-1]
        file_format = filename.rsplit(".", 1)[-1] if "." in filename else "html"
        if file_format not in {"xlsx", "xls", "csv", "pdf"} and "/publications/" not in path:
            continue
        rows.append(
            {
                "title": title,
                "format": file_format,
                "url": url,
                "source_url": source_url,
                "retrieved_at": retrieved_at,
            }
        )
    rows = list({row["url"]: row for row in rows}.values())
    if not rows:
        raise ValueError("Treasury publication page contained no recognisable resources")
    return rows


def parse_publications(html: str, source_url: str, retrieved_at: str) -> list[dict[str, str]]:
    host = urlparse(source_url).hostname
    rows = []
    for href, body in re.findall(r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', html, re.I | re.S):
        url = urljoin(source_url, unescape(href))
        title = clean(body)
        path = urlparse(url).path
        if urlparse(url).hostname != host or "/publications/" not in path or not title:
            continue
        rows.append(
            {
                "title": title,
                "url": url,
                "slug": path.rstrip("/").split("/")[-1],
                "source_url": source_url,
                "retrieved_at": retrieved_at,
                "provenance": "Treasury forecasts catalogue",
            }
        )
    rows = list({row["url"]: row for row in rows}.values())
    if not rows:
        raise ValueError("Treasury forecasts catalogue contained no publication links")
    return rows


def primary_forecast_publications(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Return primary EFU releases in the official catalogue's newest-first order."""

    title_pattern = re.compile(
        r"^(?:budget|half year|pre election) economic and fiscal update 20\d{2}$"
    )
    output = []
    for row in rows:
        normal_title = _normalise(row["title"])
        if (
            not urlparse(row["url"]).path.startswith("/publications/efu/")
            or not title_pattern.fullmatch(normal_title)
        ):
            continue
        year_match = re.search(r"\b(20\d{2})\b", row["title"])
        kind = (
            "BEFU" if normal_title.startswith("budget ")
            else "HYEFU" if normal_title.startswith("half year ")
            else "PREFU"
        )
        output.append(
            {
                **row,
                "publication_kind": kind,
                "publication_year": int(year_match.group(1)) if year_match else None,
            }
        )
    return output


# Retain the private name for callers created before the public catalogue helper.
_primary_forecast_publications = primary_forecast_publications


def resolve_publication(rows: list[dict[str, str]], query: str) -> dict[str, str] | None:
    """Resolve a unique primary EFU publication, excluding supplements and models."""

    candidates = primary_forecast_publications(rows)
    raw_query = query.strip()
    path_slug = urlparse(raw_query).path.rstrip("/").split("/")[-1] if "/" in raw_query else raw_query
    normal_query = _normalise(path_slug.replace("-", " "))
    aliases = {
        "befu": "budget economic and fiscal update",
        "hyefu": "half year economic and fiscal update",
        "prefu": "pre election economic and fiscal update",
    }
    tokens = normal_query.split()
    expanded = " ".join(aliases.get(token, token) for token in tokens)

    exact = [
        row
        for row in candidates
        if expanded in {_normalise(row["title"]), _normalise(row["slug"].replace("-", " "))}
    ]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return None

    wanted = expanded.split()
    matches = [
        row
        for row in candidates
        if wanted and all(token in _normalise(f"{row['title']} {row['slug']}").split() for token in wanted)
    ]
    return matches[0] if len(matches) == 1 else None


def select_charts_workbook(resources: list[dict[str, str]]) -> dict[str, str]:
    """Select the unique Charts and Data workbook from an EFU publication page."""

    matches = [
        row
        for row in resources
        if row["format"] == "xlsx"
        and (
            "charts and data" in _normalise(row["title"])
            or "charts data" in _normalise(urlparse(row["url"]).path)
        )
    ]
    if len(matches) != 1:
        raise ComparisonError(
            f"Expected one official Charts and Data workbook, found {len(matches)}."
        )
    return matches[0]


def _measure_definition(label: str) -> tuple[str, str]:
    normal_label = clean(label)
    if _normalise(normal_label) == "of gdp":
        raise ComparisonError("A standalone % of GDP row is ambiguous without its parent measure.")

    base = re.sub(r"\s*\([^)]*\)\s*$", "", normal_label).strip()
    lower = normal_label.casefold()
    if "$billions" in lower:
        unit = "NZD billion"
    elif "% of gdp" in lower:
        unit = "percent of GDP"
    elif "%" in normal_label or "rate" in lower:
        unit = "percent"
    else:
        raise ComparisonError(f"Cannot determine a stable unit from indicator label: {normal_label}")
    return base, unit


def parse_key_indicators(
    body: bytes,
    publication: dict[str, str],
    workbook_url: str,
    retrieved_at: str,
) -> list[dict[str, object]]:
    """Parse the compact key-indicator table used consistently across EFU workbooks."""

    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Treasury workbook could not be opened: {exc}") from exc

    if "Table 1" not in workbook.sheetnames:
        raise ValueError("Treasury Charts and Data workbook is missing the Table 1 summary")
    sheet = workbook["Table 1"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or "key economic and fiscal indicators" not in _normalise(rows[0][1] if len(rows[0]) > 1 else ""):
        raise ValueError("Treasury Table 1 is not the key economic and fiscal indicators table")

    header_index = next(
        (
            index
            for index, row in enumerate(rows[:-1])
            if any(isinstance(value, int) and 2000 <= value <= 2200 for value in row)
            and any("year ending 30 june" in _normalise(value) for value in rows[index + 1])
        ),
        None,
    )
    if header_index is None:
        raise ValueError("Treasury Table 1 year and forecast headers changed")

    years = rows[header_index]
    statuses = rows[header_index + 1]
    label_column = next(
        index for index, value in enumerate(statuses) if "year ending 30 june" in _normalise(value)
    )
    seen_measures: set[str] = set()
    records: list[dict[str, object]] = []
    for row_index, row in enumerate(rows[header_index + 2 :], start=header_index + 3):
        label = clean(str(row[label_column] or "")) if label_column < len(row) else ""
        if not label:
            continue
        try:
            measure, unit = _measure_definition(label)
        except ComparisonError as exc:
            if label == "% of GDP":
                continue
            raise exc
        measure_key = _normalise(measure)
        if measure_key in seen_measures:
            raise ComparisonError(f"Indicator definition is ambiguous within Table 1: {measure}")
        seen_measures.add(measure_key)

        for column_index, year in enumerate(years, start=1):
            if not isinstance(year, int) or not 2000 <= year <= 2200 or column_index > len(row):
                continue
            value = row[column_index - 1]
            if not isinstance(value, (int, float)):
                continue
            status = clean(str(statuses[column_index - 1] or "")).casefold()
            records.append(
                {
                    "measure": measure,
                    "measure_key": measure_key,
                    "definition": label,
                    "unit": unit,
                    "period": f"{year:04d}-06-30",
                    "period_label": str(year),
                    "period_basis": "Year ending 30 June",
                    "forecast_status": status,
                    "value": value,
                    "publication_title": publication["title"],
                    "publication_url": publication["url"],
                    "workbook_url": workbook_url,
                    "sheet": sheet.title,
                    "row": row_index,
                    "column": column_index,
                    "cell": sheet.cell(row=row_index, column=column_index).coordinate,
                    "source_url": workbook_url,
                    "retrieved_at": retrieved_at,
                    "provenance": f"{sheet.title}!{sheet.cell(row=row_index, column=column_index).coordinate}",
                }
            )
    if not records:
        raise ValueError("Treasury Table 1 contained no comparable key-indicator values")
    return records


def compare_key_indicators(
    before: list[dict[str, object]],
    after: list[dict[str, object]],
    limit: int = 100,
) -> list[dict[str, object]]:
    """Compare exact shared forecast definitions and periods without fuzzy alignment."""

    def by_measure(records: list[dict[str, object]]) -> dict[str, dict[str, object]]:
        definitions: dict[str, dict[str, object]] = {}
        for record in records:
            key = str(record["measure_key"])
            current = definitions.setdefault(
                key,
                {
                    "definition": record["definition"],
                    "unit": record["unit"],
                    "period_basis": record["period_basis"],
                },
            )
            if any(current[field] != record[field] for field in ("definition", "unit", "period_basis")):
                raise ComparisonError(f"Indicator definition is internally ambiguous: {record['measure']}")
        return definitions

    before_definitions = by_measure(before)
    after_definitions = by_measure(after)
    shared_measures = sorted(set(before_definitions) & set(after_definitions))
    for key in shared_measures:
        left = before_definitions[key]
        right = after_definitions[key]
        if left != right:
            raise ComparisonError(
                f"Refusing to align changed definition for {key}: {left!r} versus {right!r}"
            )

    def forecast_index(records: list[dict[str, object]]) -> dict[tuple[str, str], dict[str, object]]:
        output = {}
        for record in records:
            if record["forecast_status"] != "forecast" or record["measure_key"] not in shared_measures:
                continue
            key = (str(record["measure_key"]), str(record["period"]))
            if key in output:
                raise ComparisonError(f"Duplicate forecast value for {key[0]} in {key[1]}")
            output[key] = record
        return output

    left_values = forecast_index(before)
    right_values = forecast_index(after)
    shared = sorted(set(left_values) & set(right_values))
    if not shared:
        raise ComparisonError("The two publications have no exactly aligned shared forecast periods.")

    comparisons = []
    for key in shared:
        left = left_values[key]
        right = right_values[key]
        delta = round(float(right["value"]) - float(left["value"]), 12)
        comparisons.append(
            {
                "measure": left["measure"],
                "definition": left["definition"],
                "unit": left["unit"],
                "period": left["period"],
                "period_label": left["period_label"],
                "period_basis": left["period_basis"],
                "from": {field: left[field] for field in (
                    "publication_title", "publication_url", "workbook_url", "value",
                    "forecast_status", "sheet", "row", "column", "cell", "provenance",
                )},
                "to": {field: right[field] for field in (
                    "publication_title", "publication_url", "workbook_url", "value",
                    "forecast_status", "sheet", "row", "column", "cell", "provenance",
                )},
                "delta": delta,
                "delta_unit": left["unit"],
                "alignment": "exact definition, unit, period basis, and shared forecast period",
                "source_url": right["workbook_url"],
                "retrieved_at": right["retrieved_at"],
            }
        )
    return comparisons[:limit]


def filter_key_indicators(
    records: list[dict[str, object]], query: str, limit: int = 100
) -> list[dict[str, object]]:
    """Filter already-structured Table 1 values without discarding their headers."""

    needle = _normalise(query)
    if not needle:
        raise ValueError("Indicator query must not be empty")
    return [
        record
        for record in records
        if needle in _normalise(f"{record['measure']} {record['definition']}")
    ][:limit]


def search_appropriations(
    body: bytes,
    query: str,
    publication: dict[str, str],
    workbook_url: str,
    retrieved_at: str,
    limit: int = 100,
    *,
    vote_only: bool = False,
) -> list[dict[str, object]]:
    """Return one structured, provenance-complete record per published amount."""

    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Treasury workbook could not be opened: {exc}") from exc
    if "Raw Data" not in workbook.sheetnames:
        raise ValueError("Treasury expenditure workbook is missing the Raw Data sheet")

    sheet = workbook["Raw Data"]
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError("Treasury expenditure Raw Data sheet is empty") from exc
    headers = {clean(str(value or "")): index for index, value in enumerate(raw_headers)}
    required = {
        "Department", "Vote", "App ID", "Appropriation Name", "Category Name",
        "Group Type", "Appropriation or Category Type", "Restriction Type",
        "Functional Classification", "Amount $000", "Year", "Amount Type",
        "Periodicity", "Current Scope", "Portfolio Name",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(
            "Treasury expenditure Raw Data headers changed; missing: " + ", ".join(missing)
        )

    normal_query = _normalise(query)
    if vote_only and normal_query.startswith("vote "):
        normal_query = normal_query[5:].strip()
    if not normal_query:
        raise ValueError("Appropriation query must not be empty")

    def field(row: tuple[object, ...], name: str) -> object:
        index = headers[name]
        return row[index] if index < len(row) else None

    amount_column = headers["Amount $000"] + 1
    output: list[dict[str, object]] = []
    for row_number, row in enumerate(rows, start=2):
        vote = clean(str(field(row, "Vote") or ""))
        if vote_only:
            matched = _normalise(vote) == normal_query
        else:
            searchable = " ".join(
                clean(str(field(row, name) or ""))
                for name in (
                    "Department", "Vote", "Appropriation Name", "Category Name",
                    "Appropriation or Category Type", "Functional Classification",
                    "Current Scope", "Portfolio Name",
                )
            )
            matched = normal_query in _normalise(searchable)
        if not matched:
            continue

        year = field(row, "Year")
        amount = field(row, "Amount $000")
        if not isinstance(year, int) or not isinstance(amount, (int, float)):
            raise ValueError(
                f"Treasury expenditure value has an invalid year or amount at Raw Data row {row_number}"
            )
        cell = sheet.cell(row=row_number, column=amount_column).coordinate
        output.append(
            {
                "department": field(row, "Department"),
                "vote": vote,
                "app_id": field(row, "App ID"),
                "appropriation_name": field(row, "Appropriation Name"),
                "category_name": field(row, "Category Name"),
                "group_type": field(row, "Group Type"),
                "appropriation_type": field(row, "Appropriation or Category Type"),
                "restriction_type": field(row, "Restriction Type"),
                "functional_classification": field(row, "Functional Classification"),
                "current_scope": field(row, "Current Scope"),
                "portfolio_name": field(row, "Portfolio Name"),
                "value": amount,
                "unit": "NZD thousand",
                "period": f"{year:04d}-06-30",
                "period_label": str(year),
                "period_basis": "Year ending 30 June",
                "forecast_status": clean(str(field(row, "Amount Type") or "")).casefold(),
                "amount_type": field(row, "Amount Type"),
                "periodicity": field(row, "Periodicity"),
                "publication_title": publication["title"],
                "publication_url": publication["url"],
                "workbook_url": workbook_url,
                "sheet": sheet.title,
                "row": row_number,
                "column": amount_column,
                "cell": cell,
                "source_url": workbook_url,
                "retrieved_at": retrieved_at,
                "provenance": f"{sheet.title}!{cell}",
            }
        )
        if len(output) >= limit:
            break
    return output


def search_xlsx(body: bytes, query: str, workbook_url: str, retrieved_at: str, limit: int = 100):
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Treasury workbook could not be opened: {exc}") from exc
    needle = query.casefold()
    output = []
    for sheet in workbook.worksheets:
        for number, row in enumerate(sheet.iter_rows(values_only=True), 1):
            values = [value.isoformat() if hasattr(value, "isoformat") else value for value in row]
            if needle not in " ".join("" if value is None else str(value) for value in values).casefold():
                continue
            output.append(
                {
                    "publication_url": workbook_url,
                    "sheet": sheet.title,
                    "row": number,
                    "values": values,
                    "retrieved_at": retrieved_at,
                    "provenance": f"{sheet.title}!{number}",
                }
            )
            if len(output) >= limit:
                return output
    return output
