"""Parse Stats NZ Building Consents release metadata, graph CSV and workbook tables."""

from __future__ import annotations

import csv
import io
import json
import re
from html import unescape
from html.parser import HTMLParser
from urllib.parse import urljoin


SERIES = {
    "1110A1A": ("Houses", "count", "number"),
    "1121A1A": ("Apartments", "count", "number"),
    "1122A1A": ("Retirement village units", "count", "number"),
    "1129A1A": ("Townhouses, flats, and units", "count", "number"),
    "1100A1A": ("All new dwellings", "count", "number"),
    "1100A3A": ("All new dwellings", "floor_area", "thousand square metres"),
    "1100A2A": ("All new dwellings", "value", "NZD million"),
    "1000B2A": ("Residential alterations and additions", "value", "NZD million"),
    "1000C2A": ("All residential buildings", "value", "NZD million"),
    "2110C2A": ("Hostels, boarding houses, and prisons", "value", "NZD million"),
    "2120C2A": ("Hotels, motels, and short-term accommodation", "value", "NZD million"),
    "2200C2A": ("Hospitals, nursing homes, and health buildings", "value", "NZD million"),
    "2300C2A": ("Education buildings", "value", "NZD million"),
    "2400C2A": ("Social, cultural, and religious buildings", "value", "NZD million"),
    "2510C2A": ("Shops, restaurants, and bars", "value", "NZD million"),
    "2520C2A": ("Offices, administration, and public transport", "value", "NZD million"),
    "2610C2A": ("Storage buildings", "value", "NZD million"),
    "2620C2A": ("Factories and industrial buildings", "value", "NZD million"),
    "2700C2A": ("Farm buildings", "value", "NZD million"),
    "2000A3A": ("All non-residential buildings", "floor_area", "thousand square metres"),
    "2000C2A": ("All non-residential buildings", "value", "NZD million"),
    "0002C2A": ("All buildings", "value", "NZD million"),
    "3000C2A": ("Non-building construction", "value", "NZD million"),
    "0001C2A": ("All construction", "value", "NZD million"),
}


class _ReleaseParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.payload: str | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        values = dict(attrs)
        if tag == "div" and values.get("id") == "pageViewData":
            self.payload = values.get("data-value")


def _walk(value):
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from _walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from _walk(child)


def _number(value: str):
    try:
        parsed = float(value.replace(",", ""))
    except ValueError:
        return value
    return int(parsed) if parsed.is_integer() else parsed


def parse_release(html: str, source_url: str, retrieved_at: str) -> dict:
    parser = _ReleaseParser()
    parser.feed(html)
    if not parser.payload:
        raise ValueError("Stats NZ release page no longer exposes pageViewData")
    try:
        view = json.loads(parser.payload)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Stats NZ release metadata is invalid JSON: {exc}") from exc
    if not isinstance(view, dict):
        raise ValueError("Stats NZ release metadata root must be an object")

    release = view.get("DateTaxonomyTerm") or {}
    if not isinstance(release, dict):
        raise ValueError("Stats NZ release date metadata must be an object")
    documents = []
    graphs = []
    for node in _walk(view):
        if not isinstance(node, dict):
            continue
        if node.get("DocumentLink"):
            documents.append(
                {
                    "title": node.get("Title") or node.get("Name"),
                    "format": node.get("DocumentFileType"),
                    "size": node.get("DocumentSize"),
                    "url": urljoin(source_url, node["DocumentLink"]),
                    "source_url": source_url,
                    "retrieved_at": retrieved_at,
                }
            )
        if not isinstance(node.get("SeriesData"), list):
            continue
        heading = node.get("GraphHeading") or node.get("Title") or "Stats NZ series"
        for series in node["SeriesData"]:
            if not isinstance(series, dict):
                raise ValueError("Stats NZ graph series entries must be objects")
            raw = unescape(series.get("GraphCsvData") or "")
            rows = list(csv.reader(io.StringIO(raw)))
            if not rows:
                continue
            records = []
            if len(rows[0]) == 2 and re.fullmatch(r"\d{4}-\d{2}-\d{2}", rows[0][0]):
                for period, value in rows:
                    records.append(
                        {
                            "period": period,
                            "value": _number(value),
                            "series": series.get("Title"),
                            "measure": "new_dwellings",
                            "unit": "number",
                            "source_url": source_url,
                            "retrieved_at": retrieved_at,
                            "provenance": heading,
                        }
                    )
            elif len(rows) > 1:
                dimension = rows[0][0]
                categories = rows[0][1:]
                for row in rows[1:]:
                    for category, value in zip(categories, row[1:]):
                        records.append(
                            {
                                "period": row[0],
                                "dimension": dimension,
                                "category": category,
                                "value": _number(value),
                                "series": series.get("Title"),
                                "unit": "$" if "Value" in heading else "number",
                                "source_url": source_url,
                                "retrieved_at": retrieved_at,
                                "provenance": heading,
                            }
                        )
            if records:
                graphs.append({"heading": heading, "records": records})

    documents = list({item["url"]: item for item in documents}.values())
    if not documents or not graphs:
        raise ValueError("Stats NZ release schema changed: documents or graph series are missing")
    return {
        "title": view.get("Title"),
        "publication_date": release.get("PublicationDate"),
        "display_date": release.get("DateString"),
        "documents": documents,
        "graphs": graphs,
        "source_url": source_url,
        "retrieved_at": retrieved_at,
    }


def graph_records(release: dict, phrase: str) -> list[dict]:
    needle = phrase.casefold()
    for graph in release["graphs"]:
        if needle in graph["heading"].casefold():
            return graph["records"]
    raise ValueError(f"Stats NZ release no longer exposes the expected graph: {phrase}")


def parse_annual_workbook(body: bytes, workbook_url: str, retrieved_at: str) -> list[dict]:
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Stats NZ workbook could not be opened: {exc}") from exc
    records = []
    for sheet in workbook.worksheets:
        if not sheet.title.startswith("Table 1"):
            continue
        rows = list(sheet.iter_rows(values_only=True))
        code_row = next(
            (index for index, row in enumerate(rows) if any(str(value or "") in SERIES for value in row)),
            None,
        )
        if code_row is None:
            continue
        codes = {column: str(value) for column, value in enumerate(rows[code_row]) if str(value or "") in SERIES}
        annual_start = next(
            (index for index in range(code_row + 1, len(rows)) if str(rows[index][0] or "").startswith("Year ended")),
            None,
        )
        if annual_start is None:
            continue
        for row_number in range(annual_start + 1, len(rows)):
            row = rows[row_number]
            if str(row[0] or "").strip() == "Month":
                break
            if not re.fullmatch(r"\d{4}", str(row[0] or "").strip()):
                continue
            period = str(row[0]).strip()
            for column, code in codes.items():
                if column >= len(row) or not isinstance(row[column], (int, float)):
                    continue
                building_type, measure, unit = SERIES[code]
                records.append(
                    {
                        "period": period,
                        "period_context": "year ended release month",
                        "region": "New Zealand",
                        "building_type": building_type,
                        "measure": measure,
                        "value": row[column],
                        "unit": unit,
                        "series_reference": f"BLDM.SFTZ.{code}",
                        "workbook_url": workbook_url,
                        "source_url": workbook_url,
                        "retrieved_at": retrieved_at,
                        "provenance": f"{sheet.title}!{sheet.cell(row=row_number + 1, column=column + 1).coordinate}",
                        "revision_status": "latest release workbook; subject to revision",
                    }
                )
    if not records:
        raise ValueError("Stats NZ workbook schema changed: no annual Table 1 series were found")
    return records
